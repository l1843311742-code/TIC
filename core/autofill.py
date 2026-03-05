"""
智能补全与检索回填引擎 (Auto-Fill & Verification Module)
掌管了功能 [2]：利用之前入库的记忆，给那些只有一半源系统数据，却空着 SAP 字段的 Excel 行进行“填空”。
也是最核心的调度器：精确匹配 A -> 模糊查找 B -> 大模型兜底 C 的总司令部！
"""
import os
import glob
import time
import openpyxl
import chromadb
from datetime import datetime
from openpyxl.styles import Alignment
from core.config import get_logger, get_script_dir
from core.excel_utils import find_headers, map_columns
from core.llm_service import evaluate_mapping_via_llm_batch

logger = get_logger(__name__)

def auto_fill_excel(file_path: str, sheet_name: str, db_path: str, collection_name: str):
    """读取一份带有空缺目标字段的 Excel 模板，利用多级漏斗去寻找答案并写回新文件。"""
    total_start_time = time.time()
    logger.info(f"指定された不完全なフォームのスマート補完を開始します: {file_path}")
    
    # 尝试连上记忆库
    client = chromadb.PersistentClient(path=db_path)
    collection = None
    try:
        # 尝试抽出那本叫做 collection_name 的大词典
        collection = client.get_collection(name=collection_name)
    except Exception as e:
        # 万一以前根本没选过 [1] 学习功能，直接抛个错。不要紧，它照旧会坚强地运行大模型C兜底功能。
        logger.warning(f"現在、記憶データベースが空または未構築です。LLMに完全に依存して処理します。詳細: {e}")

    # 开箱 Excel
    wb = openpyxl.load_workbook(file_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    # 找坐标，拿表头，定位源系统名称
    moto_cell, saki_cell = find_headers(ws)
    if not moto_cell or not saki_cell:
        logger.error("找不到模板对齐锚点。")
        return

    # 获取全套核心列号坐标以及基准锚点
    header_row, renkei_moto_cell, renkei_saki_cell, col_src_desc, col_src_field, col_src_table, col_sap_desc, col_sap_table, col_sap_field = map_columns(ws, moto_cell, saki_cell)

    # 提取表头上部的各自系统大名称
    moto_sys_name_cell = ws.cell(row=renkei_moto_cell.row + 1, column=renkei_moto_cell.column)
    moto_sys_name = str(moto_sys_name_cell.value).strip() if moto_sys_name_cell.value else "未知Moto系"

    saki_sys_name_cell = ws.cell(row=renkei_saki_cell.row + 1, column=renkei_saki_cell.column)
    saki_sys_name = str(saki_sys_name_cell.value).strip() if saki_sys_name_cell.value else "未知Saki系"

    if None in (col_src_desc, col_src_field, col_sap_table, col_sap_field):
        logger.error("模版中残缺部分栏位")
        return

    # ======= 创建并定位：匹配来源 (A/B/C) 的标记列 =======
    col_match_source = None
    for cell in ws[header_row]:
        if cell.value in ["匹配结果来源", "匹配来源", "マッチ結果元", "マッチソース(A/B/C)", "マッチソース"]:
            col_match_source = cell.column
            break
            
    if not col_match_source:
        col_match_source = ws.max_column + 1
        ws.cell(row=header_row, column=col_match_source).value = "マッチソース"
    # ====================================================
    # 阶段零：收集全部目标 (双向自动推断方向)
    # ====================================================
    candidates = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        moto_desc = str(ws.cell(row=row_idx, column=col_src_desc).value).strip() if col_src_desc and ws.cell(row=row_idx, column=col_src_desc).value else ""
        moto_table = str(ws.cell(row=row_idx, column=col_src_table).value).strip() if col_src_table and ws.cell(row=row_idx, column=col_src_table).value else ""
        moto_field = str(ws.cell(row=row_idx, column=col_src_field).value).strip() if col_src_field and ws.cell(row=row_idx, column=col_src_field).value else ""

        saki_table = ws.cell(row=row_idx, column=col_sap_table).value
        saki_field = ws.cell(row=row_idx, column=col_sap_field).value
        saki_desc = ws.cell(row=row_idx, column=col_sap_desc).value if col_sap_desc else ""

        # 情况 A: 只有左边(Moto)有内容，右边(Saki)必须填！(正向推演)
        if moto_field and moto_desc and (not saki_table or not saki_field):
            doc_text = f"[source_system:{moto_sys_name}] [source_table:{moto_table}] [source_field:{moto_field}] [source_description:{moto_desc}]"
            candidates.append({
                "row_idx": row_idx,
                "direction": "moto_to_saki",
                "source_sys": moto_sys_name,
                "src_field": moto_field,
                "src_desc": moto_desc,
                "src_table": moto_table,
                "doc_text": doc_text
            })

        # 情况 B: 只有右边(Saki)有内容，左边(Moto)必须填！(逆向推演)
        elif saki_field and saki_desc and (not moto_table or not moto_field):
            doc_text = f"[source_system:{saki_sys_name}] [source_table:{saki_table}] [source_field:{saki_field}] [source_description:{saki_desc}]"
            candidates.append({
                "row_idx": row_idx,
                "direction": "saki_to_moto",
                "source_sys": saki_sys_name,
                "src_field": saki_field,
                "src_desc": saki_desc,
                "src_table": saki_table,
                "doc_text": doc_text
            })

    filled_count = 0
    vector_candidates = []
    
    # ===============================================================
    # 漏斗第一关：【A 级查询】精准匹配 (优化版：使用线程池并发查询)
    # 虽然ChromaDB不支持批量where条件，但可以用多线程并发执行多个查询
    # ===============================================================
    if candidates and collection:
        a_start = time.time()
        logger.info(f"レベルA 完全一致検索を開始（並列処理）: {len(candidates)} 件...")
        
        import concurrent.futures
        
        def query_exact_match(c):
            """单个精确匹配查询"""
            try:
                results = collection.query(
                    query_texts=[c["doc_text"]],
                    where={"$and": [
                        {"source_field_name": {"$eq": c["src_field"]}},
                        {"source_system_name": {"$eq": c["source_sys"]}}
                    ]},
                    n_results=1
                )
                
                if results['ids'] and results['ids'][0]:
                    return ('hit', c, results['metadatas'][0][0])
                else:
                    return ('miss', c, None)
            except Exception as e:
                logger.warning(f"完全一致検索 '{c['src_field']}' を放棄しました: {e}")
                return ('miss', c, None)
        
        # 使用线程池并发执行A级查询（IO密集型任务适合多线程）
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(query_exact_match, c) for c in candidates]
            
            for idx, future in enumerate(concurrent.futures.as_completed(futures)):
                status, c, match = future.result()
                
                if status == 'hit':
                    target_col_table = col_sap_table if c["direction"] == "moto_to_saki" else col_src_table
                    target_col_field = col_sap_field if c["direction"] == "moto_to_saki" else col_src_field
                    target_col_desc  = col_sap_desc  if c["direction"] == "moto_to_saki" else col_src_desc
                    
                    ws.cell(row=c["row_idx"], column=target_col_table).value = match['sap_table_name']
                    ws.cell(row=c["row_idx"], column=target_col_field).value = match['sap_field_name']
                    if target_col_desc:
                         ws.cell(row=c["row_idx"], column=target_col_desc).value = match.get('sap_field_desc', '')
                    ws.cell(row=c["row_idx"], column=col_match_source).value = '完全一致'
                    
                    filled_count += 1
                else:
                    vector_candidates.append(c)
                
                # 每处理10个显示一次进度
                if (idx + 1) % 10 == 0 or (idx + 1) == len(candidates):
                    logger.info(f"レベルA 進捗: {idx + 1}/{len(candidates)} 完了")
        
        a_elapsed = time.time() - a_start
        logger.info(f"レベルA 完了: {filled_count} 件ヒット (100%一致), {len(vector_candidates)} 件未ヒット (耗时: {a_elapsed:.2f}秒)")
            
    elif candidates and not collection:
        # 如果连记忆库都没有，跳过本地搜索直接给大模型兜底
        vector_candidates = candidates

    unmatched_for_llm = []
    
    # ===============================================================
    # 漏斗第二关：【批量 B 级查询】语义/维度坐标模糊相似度匹配 (Semantic/Similarity match)
    # 利用 ChromaDB 批量推理来节省计算所有语句的 Vector Embedding 耗时！
    # ===============================================================
    if vector_candidates and collection:
        b_start = time.time()
        logger.info(f"レベルB バッチ・ベクトルマッピング検索を開始: {len(vector_candidates)} 件...")
        query_texts = [c["doc_text"] for c in vector_candidates]
        
        try:
             # 获取 TOP 3
             v_results = collection.query(
                 query_texts=query_texts,
                 n_results=3
             )
             
             for i, c in enumerate(vector_candidates):
                 if v_results['ids'] and v_results['ids'][i]:
                     # 记录所有候选项的距离值，用于诊断
                     logger.debug(f"行 {c['row_idx']} [{c['doc_text'][:50]}...] 的向量候选:")
                     for j in range(len(v_results['ids'][i])):
                         meta = v_results['metadatas'][i][j]
                         dist = v_results['distances'][i][j]
                         logger.debug(f"  距离={dist:.4f} | {meta.get('sap_table_name','')}.{meta.get('sap_field_name','')} | {meta.get('sap_field_desc','')[:30]}")
                     
                     valid_matches = []
                     for j in range(len(v_results['ids'][i])):
                         distance = v_results['distances'][i][j]
                         # 必须把误差卡死在 0.10，否则像“物料”会因为和“移动类型”在数据库里余弦角度差 0.12 而被当成近义词！宁缺毋滥，把不靠谱的交给大模型。
                         if distance < 0.08: 
                             valid_matches.append((distance, v_results['metadatas'][i][j]))
                     
                     if valid_matches: 
                         # 先按距离排序（从小到大，距离越小越相似）
                         valid_matches.sort(key=lambda x: x[0])
                         
                         # 按字段名去重，保留匹配度最高的3个不同字段名
                         seen_fields = set()
                         tables = []
                         fields = []
                         descs = []
                         for d, m in valid_matches:
                             field = m.get('sap_field_name', '')
                             if field not in seen_fields:
                                 seen_fields.add(field)
                                 tables.append(m.get('sap_table_name', ''))
                                 fields.append(field)
                                 descs.append(m.get('sap_field_desc', ''))
                                 # 达到3个就停止
                                 if len(fields) >= 3:
                                     break
                         
                         target_col_table = col_sap_table if c["direction"] == "moto_to_saki" else col_src_table
                         target_col_field = col_sap_field if c["direction"] == "moto_to_saki" else col_src_field
                         target_col_desc  = col_sap_desc  if c["direction"] == "moto_to_saki" else col_src_desc

                         cell_table = ws.cell(row=c["row_idx"], column=target_col_table)
                         cell_table.value = "\n".join(tables)
                         cell_table.alignment = Alignment(wrap_text=True)
                         
                         cell_field = ws.cell(row=c["row_idx"], column=target_col_field)
                         cell_field.value = "\n".join(fields)
                         cell_field.alignment = Alignment(wrap_text=True)
                         
                         if target_col_desc: 
                             cell_desc = ws.cell(row=c["row_idx"], column=target_col_desc)
                             cell_desc.value = "\n".join(descs)
                             cell_desc.alignment = Alignment(wrap_text=True)
                             
                         ws.cell(row=c["row_idx"], column=col_match_source).value = 'ベクトル'
                         
                         filled_count += 1
                         logger.info(f"Vector Match (レベルB) ヒット {len(fields)} 件: 第一候補 {c['src_field']} -> {valid_matches[0][1].get('sap_table_name')}.{valid_matches[0][1].get('sap_field_name')} ({c['direction']})")
                     else:
                         unmatched_for_llm.append(c)
                 else:
                     unmatched_for_llm.append(c)
        except Exception as e:
             logger.warning(f"バッチ・ベクトル検索を放棄しました: {e}")
             unmatched_for_llm.extend(vector_candidates)
        
        b_elapsed = time.time() - b_start
        logger.info(f"レベルB 完了: {len(vector_candidates) - len(unmatched_for_llm)} 件ヒット, {len(unmatched_for_llm)} 件未ヒット (耗时: {b_elapsed:.2f}秒)")
        
    elif vector_candidates and not collection:
        unmatched_for_llm = vector_candidates
                
    # ===============================================================
    # 最终关底：千问大脑线上【批量】推演 (LLM Batch Request)
    # 只要攒够了没做出来的题，一次性发给老师改卷，节约无数个 HTTP 网络开销时间！
    # ===============================================================
    if unmatched_for_llm:
        llm_start = time.time()
        logger.info(f"ヒットしなかった {len(unmatched_for_llm)} 件のデータをAIモデルに一斉送信し、推論を実行します...")
        
        # O(1) 的字典返回结果
        batch_results = evaluate_mapping_via_llm_batch(unmatched_for_llm)
        llm_elapsed = time.time() - llm_start
        logger.info(f"AIモデル処理完了 (耗时: {llm_elapsed:.2f}秒, 平均: {llm_elapsed/len(unmatched_for_llm):.2f}秒/件)")
        
        for item in unmatched_for_llm:
            r_idx = item["row_idx"]
            candidates = batch_results.get(r_idx, [])
            
            if candidates and isinstance(candidates, list) and len(candidates) > 0:
                # 按字段名去重，保留前3个不同的字段名（LLM已按score排序）
                seen_fields = set()
                tables = []
                fields = []
                descs = []
                for c_item in candidates:
                    field = c_item.get('sap_field_name', '')
                    if field not in seen_fields:
                        seen_fields.add(field)
                        tables.append(c_item.get('sap_table_name', ''))
                        fields.append(field)
                        descs.append(c_item.get('sap_field_desc', ''))
                        # 达到3个就停止
                        if len(fields) >= 3:
                            break
                
                target_col_table = col_sap_table if item["direction"] == "moto_to_saki" else col_src_table
                target_col_field = col_sap_field if item["direction"] == "moto_to_saki" else col_src_field
                target_col_desc  = col_sap_desc  if item["direction"] == "moto_to_saki" else col_src_desc

                cell_table = ws.cell(row=r_idx, column=target_col_table)
                cell_table.value = "\n".join(tables)
                cell_table.alignment = Alignment(wrap_text=True)
                
                cell_field = ws.cell(row=r_idx, column=target_col_field)
                cell_field.value = "\n".join(fields)
                cell_field.alignment = Alignment(wrap_text=True)
                
                if target_col_desc: 
                    cell_desc = ws.cell(row=r_idx, column=target_col_desc)
                    cell_desc.value = "\n".join(descs)
                    cell_desc.alignment = Alignment(wrap_text=True)
                    
                ws.cell(row=r_idx, column=col_match_source).value = 'AIモデル'

                filled_count += 1
                logger.info(f"LLM バッチ補完成功: 行 {r_idx} で {len(fields)} 件の候補結果を提供しました。({item['direction']})")
            else:
                ws.cell(row=r_idx, column=col_match_source).value = '未匹配'
                logger.info(f"AIモデルも行 {r_idx} '{item['src_field']}' を推論できませんでした。完全放棄！")
    
    # 全部填写完毕！输出保存。
    if filled_count > 0:
        save_start = time.time()
        script_dir = get_script_dir()
        autofilled_folder = os.path.join(script_dir, "autofilled_output")
        os.makedirs(autofilled_folder, exist_ok=True)
        
        base, ext = os.path.splitext(os.path.basename(file_path))
        # 文件名打上当前年月日分秒的签，防止您把原模板被覆盖污染！非常安全。
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        new_filename = os.path.join(autofilled_folder, f"{base}_autofilled_{timestamp}{ext}")
        wb.save(new_filename)
        save_elapsed = time.time() - save_start
        
        total_elapsed = time.time() - total_start_time
        logger.info(f"ファイル保存完了 (耗时: {save_elapsed:.2f}秒)")
        logger.info(f"=" * 60)
        logger.info(f"大成功！合計 {filled_count} 行を補完しました！")
        logger.info(f"総処理時間: {total_elapsed:.2f}秒 (平均: {total_elapsed/filled_count:.2f}秒/行)")
        logger.info(f"保存先: {new_filename}")
        logger.info(f"=" * 60)
    else:
        logger.info("補完された行はありませんでした。（補完する空欄がない、または全レベル検索でヒットなし）。")

def process_autofill(path: str, sheet_name: str, db_path: str, collection_name: str):
    """一个小型路由器：用来决定用户丢进来的是一个文件字典文件夹结构，还是单独一个文件，然后依次下发去运行上方的 auto_fill。"""
    if os.path.isdir(path):
        excel_files = glob.glob(os.path.join(path, "*.xlsx"))
        for file in excel_files:
            if not os.path.basename(file).startswith("~") and not "_autofilled" in file:
                auto_fill_excel(file, sheet_name, db_path, collection_name)
    else:
        auto_fill_excel(path, sheet_name, db_path, collection_name)

def process_update_and_autofill(path: str, sheet_name: str, db_path: str, collection_name: str):
    """
    功能 [3] 综合模式：
    分两阶段运行。
    第一段：扫描所有表结构里，那些有SAP目标映射值的，先扒出来吃进肚子里背掉（学习ingest）；
    第二段：转身去把刚才同一批表中空着一半没SAP结果的行，调用刚刚入库背好的新鲜记忆，以及大模型尝试再填上(自动生成)。
    """
    from core.ingestion import process_ingest
    if os.path.isdir(path):
        excel_files = glob.glob(os.path.join(path, "*.xlsx"))
        # Phase 1: Ingest all available knowledge first (知识沉淀)
        for file in excel_files:
            if not os.path.basename(file).startswith("~") and not "_autofilled" in file:
                process_ingest(file, sheet_name, db_path, collection_name)
        # Phase 2: Backfill any gaps using the newly enriched DB (大反攻：回填空缺)
        for file in excel_files:
            if not os.path.basename(file).startswith("~") and not "_autofilled" in file:
                auto_fill_excel(file, sheet_name, db_path, collection_name)
    else:
        # Phase 1
        process_ingest(path, sheet_name, db_path, collection_name)
        # Phase 2
        auto_fill_excel(path, sheet_name, db_path, collection_name)
