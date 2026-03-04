"""
源数据吸纳与学习中枢 (Knowledge Ingestion Module)
掌管了功能 [1]：只读书不写字。
专门负责把用户丢进来的那些完整无缺有映射关系的 Excel 表格里，每一行的四个维度打包拼接，并且灌入到 ChromaDB 的记忆海马体。
"""
import os
import glob
import openpyxl
import chromadb
import pandas as pd
from datetime import datetime
from core.config import get_logger
from core.excel_utils import find_headers, map_columns

logger = get_logger(__name__)

def parse_excel_for_ingest(file_path: str, sheet_name: str):
    """负责将 Excel 读取到内存，找出已经全填好的行，格式化打包组装。"""
    logger.info(f"正在扫描试图用来当作学习教材的 Excel 文本: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 锁定想要找的那个名为 '項目マッピング' 的标签页
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    moto_cell, saki_cell = find_headers(ws)
    if not moto_cell or not saki_cell:
        logger.error("找不到 '連携元'。表格看起来不合规，略过。")
        return []

    # 表头上一个格子就是源系统的总名字（比如 仓库管理系统）
    source_sys_name_cell = ws.cell(row=moto_cell.row + 1, column=moto_cell.column)
    source_sys_name = str(source_sys_name_cell.value).strip() if source_sys_name_cell.value else "未知源系统"

    # 获取全套 7 个核心列号坐标
    header_row, col_src_desc, col_src_field, col_src_table, col_sap_desc, col_sap_table, col_sap_field = map_columns(ws, moto_cell, saki_cell)

    if None in (col_src_desc, col_src_field, col_sap_desc, col_sap_table, col_sap_field):
        logger.error("Excel 列不完整，忽略该学习教材！")
        return []

    extracted_data = []
    
    # 进入主题：开始遍历每一行业务数据
    for row_idx in range(header_row + 1, ws.max_row + 1):
        src_field = str(ws.cell(row=row_idx, column=col_src_field).value or "").strip()
        src_desc = str(ws.cell(row=row_idx, column=col_src_desc).value or "").strip()
        src_table = str(ws.cell(row=row_idx, column=col_src_table).value or "").strip() if col_src_table else ""
        sap_table = str(ws.cell(row=row_idx, column=col_sap_table).value or "").strip()
        sap_field = str(ws.cell(row=row_idx, column=col_sap_field).value or "").strip()
        sap_desc = str(ws.cell(row=row_idx, column=col_sap_desc).value or "").strip()
        
        # [非常关键]
        # 我们是系统在主动学习“知识库”，所以我们必须要同时挑出在：
        # 【源技术名称】、【源中文描述】 并且他们对应着确凿的 【SAP配置表】+【SAP技术名】 的那些“完成体”的行才抓下来背单词。不缺胳膊少腿的才是好学习资料！
        if src_field and src_desc and sap_table and sap_field:
            # 就是这里生成那串很长的 [source_xxx...] 给大头模型准备的特定翻译模版文本
            doc_text = f"[source_system:{source_sys_name}] [source_table:{src_table}] [source_field:{src_field}] [source_description:{src_desc}]"
            extracted_data.append({
                "id": f"mapping_{file_path}_{row_idx}", # 给这句话编一个独一无二的身份证条形码
                "text": doc_text,
                "source_system_name": source_sys_name,
                "source_table_name": src_table,
                "source_field_name": src_field,
                "source_field_desc": src_desc,
                "sap_table_name": sap_table,
                "sap_field_name": sap_field,
                "sap_field_desc": sap_desc
            })

    return extracted_data

def process_ingest(path: str, sheet_name: str, db_path: str, collection_name: str):
    """调度器：控制读取上一步取好的一堆字典结果，将他们推送入 Chroma 向量数据库完成持久记忆的写盘动作。"""
    all_documents = []
    
    # 支持丢进去整个文件夹（它会循环查出所有 .xlsx 文件拼凑），也支持单一文件
    if os.path.isdir(path):
        excel_files = glob.glob(os.path.join(path, "*.xlsx"))
        for file in excel_files:
            # 忽略打开状态占用的无形 ~ 缓存文件
            if not os.path.basename(file).startswith("~"):
                all_documents.extend(parse_excel_for_ingest(file, sheet_name))
    else:
        all_documents.extend(parse_excel_for_ingest(path, sheet_name))

    if not all_documents:
        logger.warning("扫了一大圈没发现能学的有用完整数据。")
        return

    # 连接 ChromaDB (如果库不存在会默认创建这个 sqlite 的壳)
    client = chromadb.PersistentClient(path=db_path)
    try:
        collection = client.get_collection(name=collection_name)
        
        # 拆解结构：它这行其实是在分别剥离出三个列表。
        ids = [doc['id'] for doc in all_documents] # ID纯列数组
        metadatas = [{k: v for k, v in doc.items() if k not in ['id', 'text']} for doc in all_documents] # 剔除文本后的附加结构体信息
        documents = [doc['text'] for doc in all_documents] # 给它专门塞进去的供人工智能转换的纯文字组
        
        # 触发底层大模型计算
        collection.add(ids=ids, metadatas=metadatas, documents=documents)
        logger.info(f"往现有大脑里新加入了 {len(all_documents)} 个词汇！")
    except Exception:
        # 这里的意思是：如果前面的 try 里发现没有创建名叫 collection_name 的集合空间，会出异常抛出跳转这里。
        # 在这里我们就新创立一本记忆集。hnsw:cosine代表用向量夹角算法搜查相似度（近义词算法基础）
        collection = client.create_collection(name=collection_name, metadata={"hnsw:space": "cosine"})
        
        ids = [doc['id'] for doc in all_documents]
        metadatas = [{k: v for k, v in doc.items() if k not in ['id', 'text']} for doc in all_documents]
        documents = [doc['text'] for doc in all_documents]
        
        collection.add(ids=ids, metadatas=metadatas, documents=documents)
        logger.info(f"为首次运行破土动工，新创建了集合并保存了 {len(all_documents)} 条数据。")
    
    # [贴心保障] 最后在项目内部，给你硬拷贝生成一份能被人类肉眼看懂的纯文本 CSV 大清单，当用作您的离线查看备份留底！
    from core.config import get_script_dir
    script_dir = get_script_dir()
    csv_folder = os.path.join(script_dir, "csv_output")
    os.makedirs(csv_folder, exist_ok=True)
    df = pd.DataFrame(all_documents)
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S") # 盖上 2026年3月4日... 邮编戳
    csv_filename = os.path.join(csv_folder, f"ingested_{os.path.basename(os.path.normpath(path))}_{sheet_name}_{timestamp}.csv")
    df.to_csv(csv_filename, index=False, encoding='utf-8-sig') # 写入CSV，带着utf8-sig防中文繁体乱码
    logger.info(f"为您生成的本地肉眼审查备份表已存入: {csv_filename}")
